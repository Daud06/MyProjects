from bs4 import BeautifulSoup as BS
import requests
import time
from selenium import webdriver
from selenium.webdriver.common.by import By as by
import random
from auth_data import password, username
from selenium.webdriver.common.keys import Keys
import openpyxl

headers = {
        "user-agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36",
        'accept' : 'image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8'
    }

alf = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

book = openpyxl.Workbook()
sheet = book.active
sheet.title = 'Легковые автомобили'


def get_html(url, sheet):

    r = requests.get(url, headers = headers)
    html = r.text
    soup = BS(html, 'lxml')
    child_kategory_links = soup.find('div', id = 'block-child-category').find_all('a')

    i = 1
    new_sheet = sheet
    for child_kategory in child_kategory_links:
        child_kategory_link = child_kategory.get('href')
        child_kategory_name = child_kategory.text
        print(child_kategory_link, child_kategory_name)
        if i != 1:
            new_sheet = book.create_sheet()
            new_sheet.title = child_kategory_name
        sheet = new_sheet
        child_kategory_request = requests.get(child_kategory_link, headers = headers)
        src = child_kategory_request.text
        get_link(src, child_kategory_link, sheet)
        i+=1




def get_data(html, link, str_number, sheet):
    with open('link_html.html', 'w', encoding='utf-8') as file:
        file.write(html)
    soup = BS(html, 'lxml')

    kategory = soup.find('div', id = 'crumbs').find_all('li')[3].text
    # title1 = 'Категория'

    discription = soup.find('h1').text
    # title2 = 'Описание'

    data_block = soup.find('div', id = 'ann-block')
    posting_data = data_block.find('div', id = 'ann-coords').find('div', id = 'create-date-ann').get('data-original-title').split()[0]
    title1 = data_block.find('div', id = 'ann-coords').find('div', class_ = 'title')


    city = data_block.find('div', id = 'ann-coords').find('div').find_next_siblings('div')[1].text.split()[1]
    suggest_type = data_block.find('div', id = 'ann-coords').find('div').find_next_siblings('div')[1].find('div', class_ = 'second').find('div', class_ = 'value').text.strip()
    comment = soup.find('div', id = 'ann-body').text
    username = soup.find('div', class_ = 'username').find('a').text

    try:
        options= soup.find('div', id = 'ann-options').find_all('div', class_ = 'str')
    except:
        options = []
    i = 6
    for option in options:
        title = option.find('div', class_ = 'title').text.strip()
        value = option.find('div', class_ = 'value').text.strip()
        sheet[f'{alf[i]}' + f'{str_number}'] = value
        sheet[f'A{str_number}'] = kategory
        sheet[f'B{str_number}'] = discription
        sheet[f'C{str_number}'] = link
        sheet[f'D{str_number}'] = posting_data
        sheet[f'E{str_number}'] = city
        sheet[f'F{str_number}'] = suggest_type
        sheet[f'{alf[i+1]}' + f'{str_number}'] = comment
        sheet[f'{alf[i+2]}' + f'{str_number}'] = username

        i+=1
        # print(kategory, discription, link, title, value )

    # print(kategory, discription, link,  posting_data, city, suggest_type, type, type_cargo)


def get_link(src, url, sheet):

    soup = BS(src, 'lxml')

    try:
        offers = soup.find('ul', id='change-offer-tabs').find_all('a')
    except:
        return 0
    # str_number = 2
    new_sheet = sheet
    for offer_link in offers:
        child_offer_link = offer_link.get('href')

        if url != child_offer_link:
            offer_requests = requests.get(child_offer_link, headers=headers)
            src = offer_requests.text
            soup = BS(src, 'lxml')
            new_sheet = book.create_sheet()


        title_div = soup.find_all('div', class_='ann-list-item-wrapper')
        try:
            page_count = soup.find('div', id = 'top-paginator').find_all('li')
            k = len(page_count)
            # title_div = []
            for i in range(1, k):
                if i == 1:
                    continue
                r = requests.get(url + f'/page{i}')
                page_html = r.text
                soup_page = BS(page_html, 'lxml')
                page_blocks = soup_page.find_all('div', class_ = 'ann-list-item-wrapper')
                title_div = title_div + page_blocks
            # title_div = soup.find_all('div', class_ = 'ann-list-item-wrapper')
        except:
            title_div = title_div
        # print(title_div)
        str_number = 2
        for link in title_div:
            link = link.find('h3', class_ = 'title').find('a').get('href')
            print(link)
            r = requests.get(link, headers=headers)
            html = r.text
            get_data(html, link, str_number,new_sheet)
            str_number+=1
            if str_number == 6:
                break


url1 = 'https://auto.spb.acoola.ru/category/784'
url2 = 'https://auto.spb.acoola.ru/category/82'
url3 = 'https://auto.spb.acoola.ru/category/596'
url4 = 'https://auto.spb.acoola.ru/category/83'

url = [url1, url2, url3, url4]

for i in url:
    get_html(i, sheet)


book.save('akoola.xlsx')
